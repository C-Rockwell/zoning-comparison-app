#!/usr/bin/env python3
"""
Verify Knox District Parameters.xlsx spreadsheet structure and data.
"""

from openpyxl import load_workbook
from pathlib import Path

# File path
file_path = Path("/Users/oliverseabolt/Developer/zoning-comparison-app/docs/Knox District Parameters.xlsx")

# Load workbook
try:
    wb = load_workbook(file_path)
    ws = wb.active
    print(f"✓ File opened successfully: {file_path.name}\n")
except Exception as e:
    print(f"✗ Failed to open file: {e}")
    exit(1)

# Expected district info: (col_start, code, full_name)
districts = [
    (13, "R1", "Single Family"),  # Residential baseline (checking headers only)
    (25, "RHB", "Residential Historic Building"),
    (27, "NMU", "Neighborhood Mixed Use"),
    (29, "MUCC", "Mixed Use Community Commercial"),
    (31, "HC", "Heritage Commercial"),
    (33, "BC", "Boulevard Commercial"),
    (35, "EC", "Enterprise Commercial"),
    (37, "TC", "Transit Commercial"),
]

# Expected data per district (keyed by (district_code, is_max) -> {row: value})
expected_data = {
    ("RHB", False): {5:0.5, 7:100, 14:20, 15:10, 16:10, 17:20, 21:20, 22:10, 23:10, 24:20, 52:'Y', 53:'Y', 54:'Y', 55:'Y', 58:10, 59:5, 60:10, 61:5},
    ("RHB", True): {11:'45%', 28:35, 34:35},
    ("NMU", False): {7:50, 10:'1:2', 14:5, 15:10, 16:0, 17:5, 18:0, 21:10, 22:10, 23:5, 24:10, 25:0, 30:12, 38:50, 39:40, 42:0, 43:40, 52:'N', 53:'Y', 54:'N', 55:'Y', 59:5, 61:5},
    ("NMU", True): {11:'75%', 28:35, 34:35},
    ("MUCC", False): {7:50, 16:3, 21:10, 22:5, 23:3, 24:5, 30:14, 38:60, 39:50, 43:40, 52:'N', 53:'Y', 54:'N', 55:'Y', 59:5, 61:5},
    ("MUCC", True): {11:'80%', 28:60, 34:25},
    ("HC", False): {7:100, 10:'1:2', 14:20, 15:5, 16:5, 17:10, 18:24, 21:20, 22:5, 23:5, 24:10, 25:24, 52:'Y', 53:'Y', 54:'Y', 55:'Y', 58:5, 59:5, 60:0, 61:5},
    ("HC", True): {11:'75%', 28:60, 34:25},
    ("BC", False): {7:80, 14:25, 15:10, 16:10, 17:25, 18:24, 21:25, 22:5, 23:5, 24:25, 25:24, 38:15, 39:10, 52:'Y', 53:'Y', 54:'Y', 55:'Y', 58:25, 59:10, 60:12, 61:10},
    ("BC", True): {11:'65%', 28:60, 34:25},
    ("EC", False): {5:3, 7:100, 14:10, 15:15, 16:12, 17:10, 18:24, 21:10, 22:5, 23:5, 24:10, 25:24, 30:15, 38:40, 39:30, 52:'Y', 53:'Y', 54:'Y', 55:'Y', 58:15, 59:15, 60:15, 61:15},
    ("EC", True): {11:'70%', 28:60, 34:25},
    ("TC", False): {7:25, 14:0, 15:0, 16:0, 17:5, 18:0, 21:0, 22:0, 23:0, 24:5, 25:0, 30:15, 38:70, 39:50, 52:'N', 53:'N', 54:'N', 55:'Y', 61:5},
    ("TC", True): {11:'90%', 34:25},
}

def get_cell_value(row, col):
    """Get cell value, handling Excel column indexing (1-based)."""
    cell = ws.cell(row=row, column=col)
    return cell.value

def check_headers():
    """Check residential headers (columns 13-24, row 3)."""
    print("=" * 60)
    print("CHECK 1: Residential Headers (Columns 13-24, Row 3)")
    print("=" * 60)

    passes = True
    for col in range(13, 25):
        expected = "minimum" if (col - 13) % 2 == 0 else "maximum"
        actual = get_cell_value(3, col)
        status = "✓" if actual == expected else "✗"
        if actual != expected:
            passes = False
        print(f"  Col {col}: Expected '{expected}', Got '{actual}' {status}")

    return passes

def check_commercial_headers():
    """Check commercial district headers."""
    print("\n" + "=" * 60)
    print("CHECK 2: Commercial District Headers")
    print("=" * 60)

    passes = True
    for idx, (col_start, code, full_name) in enumerate(districts[1:], 1):  # Skip R1
        print(f"\n  {code} (Columns {col_start}-{col_start+1}):")

        # Row 1: District code
        actual_code = get_cell_value(1, col_start)
        code_match = actual_code == code
        print(f"    Row 1 (code): Expected '{code}', Got '{actual_code}' {'✓' if code_match else '✗'}")
        passes = passes and code_match

        # Row 2: Full name
        actual_name = get_cell_value(2, col_start)
        name_match = actual_name == full_name
        print(f"    Row 2 (name): Expected '{full_name}', Got '{actual_name}' {'✓' if name_match else '✗'}")
        passes = passes and name_match

        # Row 3: min/max headers
        min_header = get_cell_value(3, col_start)
        max_header = get_cell_value(3, col_start + 1)
        min_match = min_header == "minimum"
        max_match = max_header == "maximum"
        print(f"    Row 3 (min col): Expected 'minimum', Got '{min_header}' {'✓' if min_match else '✗'}")
        print(f"    Row 3 (max col): Expected 'maximum', Got '{max_header}' {'✓' if max_match else '✗'}")
        passes = passes and min_match and max_match

    return passes

def check_commercial_data():
    """Check commercial district data values."""
    print("\n" + "=" * 60)
    print("CHECK 3: Commercial District Data Values")
    print("=" * 60)

    all_passes = True

    for code in ["RHB", "NMU", "MUCC", "HC", "BC", "EC", "TC"]:
        # Find column start for this district
        col_start = next(cs for cs, c, _ in districts if c == code)

        print(f"\n  {code} (Columns {col_start}-{col_start+1}):")

        # Check minimum values
        min_data = expected_data[(code, False)]
        print(f"    Minimum values: {len(min_data)} checks")
        min_passes = True
        for row, expected_val in sorted(min_data.items()):
            actual = get_cell_value(row, col_start)
            match = actual == expected_val
            if not match:
                min_passes = False
                print(f"      Row {row}: Expected {expected_val}, Got {actual} ✗")

        if min_passes:
            print(f"      All minimum values match ✓")
        all_passes = all_passes and min_passes

        # Check maximum values
        max_data = expected_data[(code, True)]
        print(f"    Maximum values: {len(max_data)} checks")
        max_passes = True
        for row, expected_val in sorted(max_data.items()):
            actual = get_cell_value(row, col_start + 1)
            match = actual == expected_val
            if not match:
                max_passes = False
                print(f"      Row {row}: Expected {expected_val}, Got {actual} ✗")

        if max_passes:
            print(f"      All maximum values match ✓")
        all_passes = all_passes and max_passes

    return all_passes

# Run all checks
check1_pass = check_headers()
check2_pass = check_commercial_headers()
check3_pass = check_commercial_data()

# Summary
print("\n" + "=" * 60)
print("VERIFICATION SUMMARY")
print("=" * 60)
print(f"Check 1 (Residential Headers):      {'PASS ✓' if check1_pass else 'FAIL ✗'}")
print(f"Check 2 (Commercial Headers):       {'PASS ✓' if check2_pass else 'FAIL ✗'}")
print(f"Check 3 (Commercial Data Values):   {'PASS ✓' if check3_pass else 'FAIL ✗'}")
print(f"\nOverall Result:                     {'PASS ✓' if (check1_pass and check2_pass and check3_pass) else 'FAIL ✗'}")
print("=" * 60)
